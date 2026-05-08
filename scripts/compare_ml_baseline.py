"""
Porownanie ML vs baseline na tych samych foldach i tym samym silniku backtestu.

Skrypt:
- wczytuje dane i buduje cechy cenowe,
- trenuje modele tylko na train,
- odpala baseline i ML na test,
- liczy wyniki przez backtests.portfolio_backtest.run_portfolio_backtest,
- wypisuje porownanie per fold oraz srednie podsumowanie.
"""

from __future__ import annotations

import os
import sys
from pathlib import Path
from typing import Any
from datetime import datetime
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from backtests.portfolio_backtest import (
    add_top_buckets,
    features_coefficient,
    final_metrics,
    run_portfolio_backtest,
)
from investment_system.features.price_features import add_price_features
from investment_system.strategies.baseline import (
    add_baseline_position,
    add_equal_weight_from_position,
)
from investment_system.strategies.ml_strategy import run_ml_strategy
from investment_system.targets.make_log_return_target import (
    add_target_excess,
    make_log_return_target,
)


import numpy as np
import pandas as pd
import yaml
from sklearn.ensemble import GradientBoostingRegressor, RandomForestRegressor
from sklearn.linear_model import LinearRegression, Ridge


HORIZON = 20
TOP_K = 15
TARGET_MODE = "absolute"  # albo "excess"
FEATURE_SET_MODE = "momentum_plus_high"
FEE_BPS = 0
BASELINE_HOLDING_PERIOD = HORIZON
ROLLING_WINDOWS = (20, 60, 120, 180, 252)
BASELINE_FAST_WINDOW = min(ROLLING_WINDOWS)
BASELINE_SIGNAL_WINDOW = max(ROLLING_WINDOWS)
FEATURE_BUCKET_COLS = []


def build_experiment_tag() -> str:
    windows_str = "-".join(str(w) for w in ROLLING_WINDOWS)
    return (
        f"h{HORIZON}"
        f"_top{TOP_K}"
        f"_fee{FEE_BPS}"
        f"_{TARGET_MODE}"
        f"_{FEATURE_SET_MODE}"
        f"_w{windows_str}"
    )


def add_excel_table(
    writer: pd.ExcelWriter,
    sheet_name: str,
    table_name: str,
    df: pd.DataFrame,
    start_row: int = 1,
    style_name: str = "TableStyleMedium2",
    show_row_stripes: bool = True,
) -> None:
    if df.empty:
        return

    ws = writer.sheets[sheet_name]
    nrows = len(df) + 1
    ncols = len(df.columns)

    end_col = get_column_letter(ncols)
    start_excel_row = start_row
    end_excel_row = start_row + nrows - 1
    table_ref = f"A{start_excel_row}:{end_col}{end_excel_row}"

    table = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=show_row_stripes,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill, Font


def add_compare_conditional_formatting(
    writer: pd.ExcelWriter,
    sheet_name: str,
    n_data_rows: int,
) -> None:
    if n_data_rows == 0:
        return

    ws = writer.sheets[sheet_name]
    row_start = 4
    row_end = 3 + n_data_rows

    # mocne, czytelne kolory
    green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")

    green_font = Font(color="006100", bold=True)
    red_font = Font(color="9C0006", bold=True)

    # D = Positive_IC_score_ratio
    # dobre >= 0.5
    ws.conditional_formatting.add(
        f"D{row_start}:D{row_end}",
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["0.5"],
            fill=green_fill,
            font=green_font,
        ),
    )
    ws.conditional_formatting.add(
        f"D{row_start}:D{row_end}",
        CellIsRule(
            operator="lessThan",
            formula=["0.5"],
            fill=red_fill,
            font=red_font,
        ),
    )

    # E:H
    # dobre >= 0
    # zle < 0
    for col in ["E", "F", "G", "H"]:
        ws.conditional_formatting.add(
            f"{col}{row_start}:{col}{row_end}",
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["0"],
                fill=green_fill,
                font=green_font,
            ),
        )
        ws.conditional_formatting.add(
            f"{col}{row_start}:{col}{row_end}",
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                fill=red_fill,
                font=red_font,
            ),
        )


for window in ROLLING_WINDOWS:
    FEATURE_BUCKET_COLS.extend(
        [
            f"distance_from_high_{window}",
            f"distance_from_min_{window}",
        ]
    )


def build_feature_cols(windows: tuple[int, ...], mode: str) -> list[str]:
    feature_cols = ["log_return"]
    if mode == "momentum_only":
        for window in windows:
            feature_cols.extend(
                [
                    f"px_log_return_mean_{window}",
                    f"px_log_return_volatility_{window}",
                ]
            )
            if window > 1:
                feature_cols.append(f"px_log_return_ratio_{window}")
    elif mode == "distance_high_only":
        for window in windows:
            feature_cols.extend(
                [
                    f"distance_from_high_{window}",
                ]
            )
    elif mode == "distance_low_only":
        for window in windows:
            feature_cols.extend(
                [
                    f"distance_from_min_{window}",
                ]
            )
    elif mode == "momentum_plus_high":
        for window in windows:
            feature_cols.extend(
                [
                    f"px_log_return_mean_{window}",
                    f"px_log_return_volatility_{window}",
                    f"distance_from_high_{window}",
                ]
            )
            if window > 1:
                feature_cols.append(f"px_log_return_ratio_{window}")
    elif mode == "momentum_plus_low":
        for window in windows:
            feature_cols.extend(
                [
                    f"px_log_return_mean_{window}",
                    f"px_log_return_volatility_{window}",
                    f"distance_from_min_{window}",
                ]
            )
            if window > 1:
                feature_cols.append(f"px_log_return_ratio_{window}")
    elif mode == "all":
        for window in windows:
            feature_cols.extend(
                [
                    f"px_log_return_mean_{window}",
                    f"px_log_return_volatility_{window}",
                    f"distance_from_high_{window}",
                    f"distance_from_min_{window}",
                ]
            )
            if window > 1:
                feature_cols.append(f"px_log_return_ratio_{window}")
    else:
        raise ValueError(f"There is no {mode} mode.")
    return feature_cols


FEATURE_COLS = build_feature_cols(ROLLING_WINDOWS, mode=FEATURE_SET_MODE)


def target_col_name(horizon: int) -> str:
    if TARGET_MODE == "absolute":
        return f"target_log_ret_{horizon}d"
    if TARGET_MODE == "excess":
        return "target_excess"
    raise ValueError("TARGET_MODE must be 'absolute' or 'excess'.")


def add_model_target(df: pd.DataFrame, horizon: int) -> pd.DataFrame:
    df = make_log_return_target(df, horizon=horizon)
    if TARGET_MODE == "excess":
        df = add_target_excess(df, horizon=horizon)
    elif TARGET_MODE != "absolute":
        raise ValueError("TARGET_MODE must be 'absolute' or 'excess'.")
    return df


def read_folds(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8") as f:
        config = yaml.safe_load(f)
    return config["folds"]


def fold_slice(df: pd.DataFrame, start: str, end: str) -> pd.DataFrame:
    start_dt = pd.to_datetime(start)
    end_dt = pd.to_datetime(end)
    return df[(df["date"] >= start_dt) & (df["date"] <= end_dt)].copy()


def run_backtest_with_metrics(
    df: pd.DataFrame,
) -> tuple[pd.DataFrame, dict[str, float]]:
    _, portfolio, metrics = run_portfolio_backtest(df, fee_bps=FEE_BPS)

    metrics["net_sharpe"] = portfolio["net_sharpe"].iloc[-1]
    metrics["max_net_drawdown_abs"] = portfolio["net_drawdown"].abs().max()
    metrics["avg_active"] = portfolio["n_active"].mean()
    metrics["avg_turnover"] = portfolio["turnover"].mean()
    return portfolio, metrics


def add_bucket_metrics(
    metrics: dict[str, float],
    df: pd.DataFrame,
    target_col: str,
    top_k: int,
) -> dict[str, float]:
    bucket_input = df.dropna(subset=["score", target_col]).copy()
    bucket_input["target"] = bucket_input[target_col]
    buckets = add_top_buckets(bucket_input, top_k)
    _, score_quality_metrics = add_score_quality_metrics(bucket_input, target_col)

    for key in (
        "bucket_n_days",
        "mean_top_buckets_return",
        "mean_bottom_buckets_return",
        "mean_bucket_spread",
    ):
        metrics[key] = buckets[key]
    metrics.update(score_quality_metrics)
    return metrics


def add_score_quality_metrics(
    df: pd.DataFrame, target_col: str
) -> tuple[pd.Series, dict]:
    ic_score = df.groupby("date")[["score", target_col]].apply(
        lambda x: x["score"].corr(x[target_col], method="spearman")
    )
    ic_score = ic_score.dropna()
    Positive_IC_score_ratio = 0
    if len(ic_score) > 0:
        Positive_IC_score_ratio = sum(np.where(ic_score > 0, 1, 0)) / len(ic_score)
    else:
        Positive_IC_score_ratio = np.nan

    IC_metrics = {
        "IC_score_mean": ic_score.mean(),
        "IC_score_median": ic_score.median(),
        "IC_score_std": ic_score.std(ddof=0),
        "IC_score_ndays": len(ic_score),
        "Positive_IC_score_ratio": Positive_IC_score_ratio,
    }

    return ic_score, IC_metrics


def feature_bucket_rows(
    df: pd.DataFrame,
    fold_name: str,
    horizon: int,
    features: tuple[str, ...],
    top_k: int,
) -> list[dict[str, Any]]:
    target_col = target_col_name(horizon)
    df_with_target = add_model_target(df, horizon=horizon)

    rows = []
    for feature in features:
        bucket_input = (
            df_with_target.replace([np.inf, -np.inf], np.nan)
            .dropna(subset=[feature, target_col])
            .copy()
        )
        bucket_input["score"] = bucket_input[feature]
        bucket_input["target"] = bucket_input[target_col]

        buckets = add_top_buckets(bucket_input, top_k)
        _, score_quality_metrics = add_score_quality_metrics(bucket_input, target_col)
        rows.append(
            {
                "fold": fold_name,
                "feature_name": feature,
                "top_k": top_k,
                "bucket_n_days": buckets["bucket_n_days"],
                "mean_top_buckets_return": buckets["mean_top_buckets_return"],
                "mean_bottom_buckets_return": buckets["mean_bottom_buckets_return"],
                "mean_bucket_spread": buckets["mean_bucket_spread"],
                **score_quality_metrics,
            }
        )

    return rows


def feature_coefficient_rows(
    df: pd.DataFrame,
    fold_name: str,
    horizon: int,
    features: list[str],
) -> list[dict[str, Any]]:
    target_col = target_col_name(horizon)
    df_with_target = add_model_target(df, horizon=horizon)
    feature_input = (
        df_with_target.replace([np.inf, -np.inf], np.nan)
        .dropna(subset=[*features, target_col])
        .copy()
    )
    feature_input["target"] = feature_input[target_col]

    coefficients = features_coefficient(feature_input, features)
    if isinstance(coefficients, pd.DataFrame):
        rows = coefficients.to_dict("records")
        return [
            {
                "fold": fold_name,
                "horizon": horizon,
                **row,
            }
            for row in rows
        ]

    rows = []
    for feature in features:
        rows.append(
            {
                "fold": fold_name,
                "horizon": horizon,
                "feature_name": feature,
                "IC_Mean": coefficients.get(f"{feature}_IC_Mean"),
                "IC_Median": coefficients.get(f"{feature}_IC_Median"),
                "IC_Std": coefficients.get(f"{feature}_IC_Std"),
                "IC_N_Days": coefficients.get(f"{feature}_IC_N_Days"),
                "Positive_IC_Rate": coefficients.get(f"{feature}_Positive_IC_Rate"),
            }
        )
    return rows


def add_baseline_signal_with_train_threshold(
    df_all: pd.DataFrame,
    train_end: str,
    top_k: int,
    holding_period: int,
) -> pd.DataFrame:
    """
    Baseline bez przecieku z testu: prog 70% jest liczony tylko na train.
    Decyzja jest odswiezana co holding_period dni, a wagi i wykonanie
    ida pozniej przez wspolny portfolio_backtest.
    """
    if holding_period <= 0:
        raise ValueError("holding_period must be greater than 0.")

    df_all = df_all.copy()

    train_end_dt = pd.to_datetime(train_end)
    fast_mean_col = f"px_log_return_mean_{BASELINE_FAST_WINDOW}"
    signal_mean_col = f"px_log_return_mean_{BASELINE_SIGNAL_WINDOW}"
    signal_volatility_col = f"px_log_return_volatility_{BASELINE_SIGNAL_WINDOW}"
    signal_ratio_col = f"px_log_return_ratio_{BASELINE_SIGNAL_WINDOW}"

    df_all["score"] = (
        df_all[signal_mean_col]
        * df_all[fast_mean_col]
        / (df_all[signal_volatility_col] + 1e-12)
    )
    df_all["rank"] = df_all.groupby("date")["score"].rank(ascending=False)

    ratio = df_all[signal_ratio_col].replace([np.inf, -np.inf], np.nan)
    threshold = ratio[df_all["date"] <= train_end_dt].quantile(0.70)

    cond = (
        (df_all[signal_mean_col] > 0)
        & (df_all[fast_mean_col] > 0)
        & (ratio > threshold)
        & (df_all["rank"] <= top_k)
    )

    dates = pd.Series(df_all["date"].drop_duplicates().sort_values().to_numpy())
    rebalance_dates = set(dates.iloc[::holding_period])
    df_all["rebalance_day"] = df_all["date"].isin(rebalance_dates)

    df_all["signal_decision"] = np.where(
        df_all["rebalance_day"],
        np.where(cond, 1, 0),
        np.nan,
    )
    df_all = df_all.sort_values(["symbol", "date"]).reset_index(drop=True)
    df_all["signal_hold"] = (
        df_all.groupby("symbol")["signal_decision"].ffill().fillna(0)
    )
    df_all = add_baseline_position(df_all)
    df_all = add_equal_weight_from_position(df_all)
    return df_all


def run_baseline_on_fold(
    df_train: pd.DataFrame,
    df_test: pd.DataFrame,
    train_end: str,
    top_k: int,
    holding_period: int,
) -> dict[str, float]:
    target_col = target_col_name(HORIZON)
    df_all = pd.concat([df_train, df_test], ignore_index=True)
    df_all = add_model_target(df_all, horizon=HORIZON)
    df_all = df_all.sort_values(["symbol", "date"]).reset_index(drop=True)
    df_all = add_baseline_signal_with_train_threshold(
        df_all,
        train_end,
        top_k,
        holding_period,
    )

    test_start = df_test["date"].min()
    df_test_result = df_all[df_all["date"] >= test_start].copy()
    _, metrics = run_backtest_with_metrics(df_test_result)
    metrics = add_bucket_metrics(metrics, df_test_result, target_col, top_k)
    return metrics


def run_ml_on_fold(
    df_train: pd.DataFrame,
    df_test: pd.DataFrame,
    model_class: type,
    model_kwargs: dict[str, Any],
    horizon: int,
    top_k: int,
) -> dict[str, float]:
    target_col = target_col_name(horizon)
    train_with_target = add_model_target(df_train, horizon=horizon)
    train_clean = (
        train_with_target[FEATURE_COLS + [target_col]]
        .replace([np.inf, -np.inf], np.nan)
        .dropna()
    )

    if train_clean.empty:
        raise ValueError("Train fold is empty after dropping missing features/target.")

    model = model_class(**model_kwargs)
    model.fit(train_clean[FEATURE_COLS], train_clean[target_col])

    df_test_with_target = add_model_target(df_test, horizon=horizon)
    df_test_scored = (
        df_test_with_target.replace([np.inf, -np.inf], np.nan)
        .dropna(subset=FEATURE_COLS)
        .copy()
    )
    if df_test_scored.empty:
        raise ValueError("Test fold is empty after dropping missing features.")

    df_test_scored["score"] = model.predict(df_test_scored[FEATURE_COLS])
    df_test_scored = run_ml_strategy(df_test_scored, horizon, top_k)

    _, metrics = run_backtest_with_metrics(df_test_scored)
    metrics = add_bucket_metrics(metrics, df_test_scored, target_col, top_k)
    return metrics


def fmt(value: Any, kind: str = "float") -> str:
    if pd.isna(value):
        return "nan"
    if kind == "int":
        return f"{int(value)}"
    if kind == "pct":
        return f"{value * 100:,.1f}%"
    return f"{value:,.3f}"


def add_vs_baseline_metrics(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_fold = {row["fold"]: row for row in results if row["model"] == "Baseline"}

    enriched = []
    for row in results:
        baseline = by_fold[row["fold"]]
        enriched.append(
            {
                **row,
                "vs_baseline_equity": (
                    row["final_net_equity"] - baseline["final_net_equity"]
                ),
                "vs_baseline_sharpe": row["net_sharpe"] - baseline["net_sharpe"],
            }
        )
    return enriched


def print_table(rows: list[dict[str, Any]], columns: list[tuple[str, str]]) -> None:
    prepared = []
    for row in rows:
        prepared.append([str(row.get(key, "")) for key, _ in columns])

    widths = []
    for idx, (_, label) in enumerate(columns):
        values = [line[idx] for line in prepared]
        widths.append(max([len(label), *map(len, values)]))

    header = " | ".join(
        label.ljust(widths[idx]) for idx, (_, label) in enumerate(columns)
    )
    sep = "-+-".join("-" * width for width in widths)
    print(header)
    print(sep)
    for line in prepared:
        print(" | ".join(line[idx].ljust(widths[idx]) for idx in range(len(columns))))


def save_results(
    all_metrics: list[dict[str, Any]],
    display_rows: list[dict[str, str]],
    summary_rows: list[dict[str, str]],
    feature_ic_rows: list[dict[str, Any]],
    feature_bucket_rows_: list[dict[str, Any]],
) -> Path:
    out_dir = ROOT / "results" / "backtests"
    out_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    experiment_tag = build_experiment_tag()
    out_path = out_dir / f"ml_baseline_comparison_{experiment_tag}_{timestamp}.xlsx"

    raw_metrics_df = pd.DataFrame(all_metrics)
    folds_df = pd.DataFrame(display_rows)
    summary_df = pd.DataFrame(summary_rows)
    feature_ic_df = pd.DataFrame(feature_ic_rows)
    feature_buckets_df = pd.DataFrame(feature_bucket_rows_)

    compare_view_df = (
        raw_metrics_df.loc[
            raw_metrics_df["model"] != "Baseline",
            [
                "fold",
                "model",
                "final_net_equity",
                "Positive_IC_score_ratio",
                "vs_baseline_equity",
                "vs_baseline_sharpe",
                "IC_score_mean",
                "mean_bucket_spread",
            ],
        ]
        .rename(
            columns={
                "fold": "FOLD",
                "model": "MODEL",
                "final_net_equity": "FINAL_NET_EQUITY",
                "Positive_IC_score_ratio": "Positive_IC_score_ratio",
                "vs_baseline_equity": "vs_baseline_equity",
                "vs_baseline_sharpe": "vs_baseline_SHARPE",
                "IC_score_mean": "IC_score_mean",
                "mean_bucket_spread": "MEAN_BUCKET_SPREAD",
            }
        )
        .sort_values(["FOLD", "MODEL"])
        .reset_index(drop=True)
    )

    config_df = pd.DataFrame(
        [
            {"parameter": "HORIZON", "value": HORIZON},
            {"parameter": "TOP_K", "value": TOP_K},
            {"parameter": "FEE_BPS", "value": FEE_BPS},
            {
                "parameter": "BASELINE_HOLDING_PERIOD",
                "value": BASELINE_HOLDING_PERIOD,
            },
            {"parameter": "TARGET_MODE", "value": TARGET_MODE},
            {"parameter": "FEATURE_SET_MODE", "value": FEATURE_SET_MODE},
            {
                "parameter": "ROLLING_WINDOWS",
                "value": ",".join(map(str, ROLLING_WINDOWS)),
            },
            {"parameter": "BASELINE_FAST_WINDOW", "value": BASELINE_FAST_WINDOW},
            {"parameter": "BASELINE_SIGNAL_WINDOW", "value": BASELINE_SIGNAL_WINDOW},
            {"parameter": "FEATURE_COLS", "value": ", ".join(FEATURE_COLS)},
            {
                "parameter": "FEATURE_BUCKET_COLS",
                "value": ", ".join(FEATURE_BUCKET_COLS),
            },
        ]
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        raw_metrics_df.to_excel(writer, sheet_name="raw_metrics", index=False)
        folds_df.to_excel(writer, sheet_name="folds", index=False)
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        feature_ic_df.to_excel(writer, sheet_name="feature_ic", index=False)
        feature_buckets_df.to_excel(writer, sheet_name="feature_buckets", index=False)
        config_df.to_excel(writer, sheet_name="config", index=False)

        compare_view_df.to_excel(
            writer,
            sheet_name="compare_view",
            index=False,
            startrow=2,
        )

        add_excel_table(writer, "raw_metrics", "tbl_raw_metrics", raw_metrics_df)
        add_excel_table(writer, "folds", "tbl_folds", folds_df)
        add_excel_table(writer, "summary", "tbl_summary", summary_df)
        add_excel_table(writer, "feature_ic", "tbl_feature_ic", feature_ic_df)
        add_excel_table(
            writer,
            "feature_buckets",
            "tbl_feature_buckets",
            feature_buckets_df,
        )
        add_excel_table(writer, "config", "tbl_config", config_df)
        add_excel_table(
            writer,
            "compare_view",
            "tbl_compare_view",
            compare_view_df,
            start_row=3,
            style_name="TableStyleLight1",
            show_row_stripes=False,
        )

        add_compare_conditional_formatting(
            writer,
            "compare_view",
            len(compare_view_df),
        )

    return out_path


def result_row(
    fold_name: str, model_name: str, metrics: dict[str, float]
) -> dict[str, str]:
    return {
        "fold": fold_name,
        "model": "BASELINE" if model_name == "Baseline" else model_name,
        "equity": fmt(metrics["final_net_equity"]),
        "vs_equity": fmt(metrics["vs_baseline_equity"]),
        "sharpe": fmt(metrics["net_sharpe"]),
        "vs_sharpe": fmt(metrics["vs_baseline_sharpe"]),
        "max_dd": fmt(metrics["max_net_drawdown_abs"], "pct"),
        "avg_active": fmt(metrics["avg_active"]),
        "turnover": fmt(metrics["avg_turnover"]),
        "trade_rate": fmt(metrics["trade_rate"], "pct"),
        "bucket_top": fmt(metrics["mean_top_buckets_return"], "pct"),
        "bucket_bottom": fmt(metrics["mean_bottom_buckets_return"], "pct"),
        "bucket_spread": fmt(metrics["mean_bucket_spread"], "pct"),
        "bucket_days": fmt(metrics["bucket_n_days"], "int"),
        "score_ic": fmt(metrics["IC_score_mean"]),
        "score_ic_pos": fmt(metrics["Positive_IC_score_ratio"], "pct"),
        "score_ic_days": fmt(metrics["IC_score_ndays"], "int"),
        "days": fmt(metrics["n_days"], "int"),
    }


def summarize_results(results: list[dict[str, Any]]) -> list[dict[str, str]]:
    raw_summary = {}
    for model_name, group in pd.DataFrame(results).groupby("model"):
        raw_summary[model_name] = {
            "chained_equity": group["final_net_equity"].prod(),
            "avg_equity": group["final_net_equity"].mean(),
            "avg_sharpe": group["net_sharpe"].mean(),
            "avg_max_dd": group["max_net_drawdown_abs"].mean(),
            "avg_active": group["avg_active"].mean(),
            "avg_turnover": group["avg_turnover"].mean(),
            "avg_trade_rate": group["trade_rate"].mean(),
            "avg_bucket_top": group["mean_top_buckets_return"].mean(),
            "avg_bucket_bottom": group["mean_bottom_buckets_return"].mean(),
            "avg_bucket_spread": group["mean_bucket_spread"].mean(),
            "avg_bucket_days": group["bucket_n_days"].mean(),
            "avg_score_ic": group["IC_score_mean"].mean(),
            "avg_score_ic_positive_rate": group["Positive_IC_score_ratio"].mean(),
            "avg_score_ic_days": group["IC_score_ndays"].mean(),
            "avg_vs_baseline_equity": group["vs_baseline_equity"].mean(),
            "avg_vs_baseline_sharpe": group["vs_baseline_sharpe"].mean(),
        }

    baseline_chained_equity = raw_summary["Baseline"]["chained_equity"]

    summary = []
    for model_name, values in raw_summary.items():
        summary.append(
            {
                "model": "BASELINE" if model_name == "Baseline" else model_name,
                "chained_equity": fmt(values["chained_equity"]),
                "vs_chained_equity": fmt(
                    values["chained_equity"] - baseline_chained_equity
                ),
                "equity": fmt(values["avg_equity"]),
                "vs_equity": fmt(values["avg_vs_baseline_equity"]),
                "sharpe": fmt(values["avg_sharpe"]),
                "vs_sharpe": fmt(values["avg_vs_baseline_sharpe"]),
                "max_dd": fmt(values["avg_max_dd"], "pct"),
                "avg_active": fmt(values["avg_active"]),
                "turnover": fmt(values["avg_turnover"]),
                "trade_rate": fmt(values["avg_trade_rate"], "pct"),
                "bucket_top": fmt(values["avg_bucket_top"], "pct"),
                "bucket_bottom": fmt(values["avg_bucket_bottom"], "pct"),
                "bucket_spread": fmt(values["avg_bucket_spread"], "pct"),
                "bucket_days": fmt(values["avg_bucket_days"]),
                "score_ic": fmt(values["avg_score_ic"]),
                "score_ic_pos": fmt(values["avg_score_ic_positive_rate"], "pct"),
                "score_ic_days": fmt(values["avg_score_ic_days"]),
            }
        )
    return sorted(
        summary, key=lambda row: float(row["sharpe"].replace(",", "")), reverse=True
    )


def feature_bucket_display_rows(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    return [
        {
            "fold": row["fold"],
            "feature": row["feature_name"],
            "top": fmt(row["mean_top_buckets_return"], "pct"),
            "bottom": fmt(row["mean_bottom_buckets_return"], "pct"),
            "spread": fmt(row["mean_bucket_spread"], "pct"),
            "score_ic": fmt(row["IC_score_mean"]),
            "ic_pos": fmt(row["Positive_IC_score_ratio"], "pct"),
            "days": fmt(row["bucket_n_days"], "int"),
        }
        for row in rows
    ]


def summarize_feature_buckets(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    summary = []
    for feature_name, group in pd.DataFrame(rows).groupby("feature_name"):
        summary.append(
            {
                "feature": feature_name,
                "top": fmt(group["mean_top_buckets_return"].mean(), "pct"),
                "bottom": fmt(group["mean_bottom_buckets_return"].mean(), "pct"),
                "spread": fmt(group["mean_bucket_spread"].mean(), "pct"),
                "score_ic": fmt(group["IC_score_mean"].mean()),
                "ic_pos": fmt(group["Positive_IC_score_ratio"].mean(), "pct"),
                "days": fmt(group["bucket_n_days"].mean()),
            }
        )
    return summary


def main() -> None:
    print("=" * 88)
    print("ML vs BASELINE - wspolny backtest portfelowy")
    print("=" * 88)
    print(
        f"horizon={HORIZON}, top_k={TOP_K}, fee_bps={FEE_BPS}, "
        f"baseline_holding_period={BASELINE_HOLDING_PERIOD}, "
        f"target_mode={TARGET_MODE}, "
        f"feature_set_mode={FEATURE_SET_MODE}"
    )

    folds = read_folds(ROOT / "configs" / "folds.yaml")

    prices_path = ROOT / "data" / "raw" / "prices.parquet"
    df = pd.read_parquet(prices_path)
    df = add_price_features(df, ROLLING_WINDOWS)

    models = {
        "Baseline": None,
        "Ridge": (Ridge, {"alpha": 1.0}),
        "Linear": (LinearRegression, {}),
        "GradientBoosting": (
            GradientBoostingRegressor,
            {"n_estimators": 50, "max_depth": 3, "random_state": 42},
        ),
        "RandomForest": (
            RandomForestRegressor,
            {"n_estimators": 100, "max_depth": 6, "random_state": 42, "n_jobs": -1},
        ),
    }

    all_metrics: list[dict[str, Any]] = []
    feature_ic_rows: list[dict[str, Any]] = []
    feature_bucket_rows_: list[dict[str, Any]] = []

    for idx, fold in enumerate(folds, start=1):
        fold_name = f"Fold {idx}"
        df_train = fold_slice(df, fold["train_start"], fold["train_end"])
        df_test = fold_slice(df, fold["test_start"], fold["test_end"])

        print()
        print(
            f"{fold_name}: train {fold['train_start']}..{fold['train_end']} | "
            f"test {fold['test_start']}..{fold['test_end']} | "
            f"rows train={len(df_train):,}, test={len(df_test):,}"
        )
        feature_ic_rows.extend(
            feature_coefficient_rows(df_test, fold_name, HORIZON, FEATURE_COLS)
        )
        feature_bucket_rows_.extend(
            feature_bucket_rows(
                df_test,
                fold_name,
                HORIZON,
                FEATURE_BUCKET_COLS,
                TOP_K,
            )
        )

        for model_name, model_spec in models.items():
            if model_spec is None:
                metrics = run_baseline_on_fold(
                    df_train,
                    df_test,
                    train_end=fold["train_end"],
                    top_k=TOP_K,
                    holding_period=BASELINE_HOLDING_PERIOD,
                )
            else:
                model_class, model_kwargs = model_spec
                metrics = run_ml_on_fold(
                    df_train,
                    df_test,
                    model_class,
                    model_kwargs,
                    horizon=HORIZON,
                    top_k=TOP_K,
                )

            metrics = {"fold": fold_name, "model": model_name, **metrics}
            all_metrics.append(metrics)

    all_metrics = add_vs_baseline_metrics(all_metrics)
    display_rows = [result_row(row["fold"], row["model"], row) for row in all_metrics]
    summary_rows = summarize_results(all_metrics)
    results_path = save_results(
        all_metrics,
        display_rows,
        summary_rows,
        feature_ic_rows,
        feature_bucket_rows_,
    )

    columns = [
        ("fold", "Fold"),
        ("model", "Model"),
        ("equity", "Net equity"),
        ("vs_equity", "vs BASE eq"),
        ("sharpe", "Net Sharpe"),
        ("vs_sharpe", "vs BASE Sh"),
        ("max_dd", "Max DD"),
        ("avg_active", "Avg active"),
        ("turnover", "Turnover"),
        ("trade_rate", "Trade rate"),
        ("bucket_top", "Top bucket"),
        ("bucket_bottom", "Bottom bucket"),
        ("bucket_spread", "Bucket spread"),
        ("bucket_days", "Bucket days"),
        ("score_ic", "Score IC"),
        ("score_ic_pos", "IC > 0"),
        ("score_ic_days", "IC days"),
        ("days", "Days"),
    ]

    print()
    print("=" * 88)
    print("Baseline - wyniki po foldach")
    print("=" * 88)
    print_table(
        [row for row in display_rows if row["model"] == "BASELINE"],
        columns,
    )

    print()
    print("=" * 88)
    print("Wyniki po foldach")
    print("=" * 88)
    print_table(display_rows, columns)

    print()
    print("=" * 88)
    print("Srednie po foldach")
    print("=" * 88)
    print_table(
        summary_rows,
        [
            ("model", "Model"),
            ("chained_equity", "Chained eq"),
            ("vs_chained_equity", "vs BASE chain"),
            ("equity", "Avg net equity"),
            ("vs_equity", "vs BASE eq"),
            ("sharpe", "Avg net Sharpe"),
            ("vs_sharpe", "vs BASE Sh"),
            ("max_dd", "Avg max DD"),
            ("avg_active", "Avg active"),
            ("turnover", "Avg turnover"),
            ("trade_rate", "Avg trade rate"),
            ("bucket_top", "Avg top bucket"),
            ("bucket_bottom", "Avg bottom bucket"),
            ("bucket_spread", "Avg bucket spread"),
            ("bucket_days", "Avg bucket days"),
            ("score_ic", "Avg score IC"),
            ("score_ic_pos", "Avg IC > 0"),
            ("score_ic_days", "Avg IC days"),
        ],
    )

    feature_bucket_columns = [
        ("fold", "Fold"),
        ("feature", "Feature"),
        ("top", "Top bucket"),
        ("bottom", "Bottom bucket"),
        ("spread", "Bucket spread"),
        ("score_ic", "Feature IC"),
        ("ic_pos", "IC > 0"),
        ("days", "Days"),
    ]

    print()
    print("=" * 88)
    print("Bucket return dla cech")
    print("=" * 88)
    print_table(
        feature_bucket_display_rows(feature_bucket_rows_), feature_bucket_columns
    )

    print()
    print("=" * 88)
    print("Srednie bucket return dla cech")
    print("=" * 88)
    print_table(
        summarize_feature_buckets(feature_bucket_rows_),
        [
            ("feature", "Feature"),
            ("top", "Avg top bucket"),
            ("bottom", "Avg bottom bucket"),
            ("spread", "Avg bucket spread"),
            ("score_ic", "Avg feature IC"),
            ("ic_pos", "Avg IC > 0"),
            ("days", "Avg days"),
        ],
    )
    print()
    print(f"Zapisano wyniki: {results_path}")


if __name__ == "__main__":
    main()
